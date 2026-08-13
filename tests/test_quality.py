"""数据质量模块单元测试。"""
from datetime import datetime, timedelta

from quality import (
    EXPORT_COLUMNS,
    ISSUE_FIELDS,
    STALE_DAYS,
    build_csv,
    build_xlsx,
    classify_row,
    collect_title_fixes,
    is_deletable,
    summarize,
)


def make_row(**over):
    row = {
        'video_id': '1',
        'video_title': '标题',
        'video_desc': '描述',
        'author_name': '作者',
        'author_id': 'A1',
        'publish_time': datetime(2026, 1, 1),
        'like_count': 1,
        'comment_count': 1,
        'share_count': 1,
        'play_count': 0,
        'video_url': 'u',
        'cover_url': 'c',
        'crawl_time': datetime(2026, 5, 20),
        'update_time': datetime(2026, 8, 10),
    }
    row.update(over)
    return row


def test_classify_empty_record():
    row = make_row(video_title='', author_name='')
    assert 'empty' in classify_row(row)


def test_classify_placeholder():
    row = make_row(video_title='在抖音记录美好生活 - 抖音')
    assert 'placeholder' in classify_row(row)


def test_classify_stale():
    row = make_row(update_time=datetime.now() - timedelta(days=STALE_DAYS + 1))
    assert 'stale' in classify_row(row)


def test_classify_missing_author_only():
    row = make_row(author_name='')
    issues = classify_row(row)
    assert 'missing_author' in issues
    assert 'empty' not in issues


def test_classify_clean_row_has_no_issues():
    assert classify_row(make_row()) == []


def test_summarize_counts():
    rows = [
        make_row(video_id='1'),
        make_row(video_id='2', video_title='', author_name=''),
    ]
    summary = summarize(rows)
    assert summary['total'] == 2
    assert summary['distinct_video_ids'] == 2
    assert summary['authors'] == 1
    assert summary['issue_counts']['empty'] == 1


def test_collect_title_fixes_only_whitespace_changes():
    rows = [
        make_row(video_id='1', video_title='  标题 \n第二行 '),
        make_row(video_id='2'),
    ]
    fixes = collect_title_fixes(rows)
    assert fixes == [('1', '标题 第二行')]


def test_is_deletable_uses_current_row_not_report_snapshot():
    row_empty = make_row(video_title='', author_name='')
    assert is_deletable(row_empty) is True
    # 报告生成后该行被补全，删除时应拒绝
    row_fixed = make_row(video_title='已补全', author_name='作者', like_count=5)
    assert is_deletable(row_fixed) is False


def test_is_deletable_true_for_stale_row():
    row = make_row(update_time=datetime.now() - timedelta(days=STALE_DAYS + 1))
    assert is_deletable(row) is True


def test_build_csv_escapes_special_characters_and_adds_bom():
    row = make_row(video_title='标题,带逗号\n换行"引号"')
    text = build_csv([row])
    assert text.startswith('\ufeff')
    assert '"标题,带逗号\n换行""引号"""' in text


def test_build_csv_includes_header():
    text = build_csv([make_row()])
    header = text.splitlines()[0]
    assert 'video_id' in header
    assert 'video_title' in header
    assert 'update_time' in header


def test_build_xlsx_returns_valid_workbook():
    rows = [
        make_row(video_id='1', video_title='标题,带逗号"引号"'),
        make_row(video_id='2'),
    ]
    content = build_xlsx(rows)
    assert content[:2] == b'PK'  # xlsx 本质是 zip 包

    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    assert ws.max_row == 3  # 表头 + 2 行数据
    assert ws.cell(1, 1).value == 'video_id'
    assert ws.cell(2, 1).value == '1'
    assert ws.cell(2, 2).value == '标题,带逗号"引号"'


def test_export_and_issue_fields_include_collect_count():
    assert 'collect_count' in EXPORT_COLUMNS
    assert 'collect_count' in ISSUE_FIELDS
    header = build_csv([make_row(collect_count=66)]).splitlines()[0]
    assert 'collect_count' in header
